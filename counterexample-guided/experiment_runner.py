"""
CTX-verify Experiment Runner
Automated verification and repair experiment using DeepSeek Reasoner API
"""

import os
import sys
import json
import time
import subprocess
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
import re
import shutil
from openai import OpenAI
from config import *

class ExperimentRunner:
    def __init__(self, api_key, base_dir=BASE_DIR):
        self.api_key = api_key
        self.base_dir = Path(base_dir)
        self.timestamp = datetime.now().strftime("%m%d%H%M")
        self.results_dir = Path(f"{RESULTS_DIR_PREFIX}-{self.timestamp}")
        self.excel_data = []
        self.api_base_url = API_BASE_URL
        self.model_name = MODEL_NAME
        
        # Initialize DeepSeek client
        self.client = OpenAI(
            api_key=api_key,
            base_url="https://api.deepseek.com"
        )
        
        # Create results directory
        self.results_dir.mkdir(exist_ok=True)
        
    def call_deepseek_reasoner(self, prompt, max_retries=MAX_API_RETRIES):
        """Call DeepSeek Reasoner API with retry logic"""
        for attempt in range(max_retries):
            try:
                print(f"    API call attempt {attempt + 1} - Model: {self.model_name}")
                print(f"    Prompt length: {len(prompt)} characters")
                
                response = self.client.chat.completions.create(
                    model=self.model_name,
                    messages=[
                        {"role": "user", "content": prompt}
                    ],
                    stream=False
                )
                
                print(f"    Response status: Success")
                
                # Check if response has the expected structure
                if hasattr(response, 'choices') and len(response.choices) > 0:
                    content = response.choices[0].message.content
                    print(f"    Response content length: {len(content)} characters")
                    
                    # Check if content is empty or very short
                    if not content or len(content.strip()) < 10:
                        print(f"    WARNING: API returned very short/empty content: '{content[:100]}...'")
                        if attempt < max_retries - 1:
                            print(f"    Retrying... (attempt {attempt + 1}/{max_retries})")
                            time.sleep(2 ** attempt)
                            continue
                    
                    return content
                else:
                    print(f"    ERROR: Unexpected response structure: {response}")
                    if attempt < max_retries - 1:
                        time.sleep(2 ** attempt)
                        continue
                    
            except Exception as e:
                print(f"    API call attempt {attempt + 1} failed: {e}")
                if attempt < max_retries - 1:
                    time.sleep(2 ** attempt)
                    continue
                    
        print(f"    All {max_retries} API call attempts failed")
        return None
    
    def run_verus(self, file_path):
        """Run verus verification on a file"""
        try:
            result = subprocess.run(
                ["verus", str(file_path)],
                capture_output=True,
                text=True,
                timeout=VERUS_TIMEOUT
            )
            return result.stdout, result.stderr, result.returncode
        except subprocess.TimeoutExpired:
            return "", "Verification timeout", 1
        except Exception as e:
            return "", f"Verification error: {e}", 1
    
    def check_verus_success(self, stdout, stderr):
        """Check if verus verification was successful"""
        combined_output = stdout + stderr
        return "0 errors" in combined_output
    
    def extract_python_code(self, response):
        """Extract Python code from API response"""
        # Try to find Python code blocks
        code_patterns = [
            r'```python\s*(.*?)\s*```',
            r'```\s*(.*?)\s*```',
            r'```\s*(.*?)\s*```',
        ]
        
        for pattern in code_patterns:
            matches = re.findall(pattern, response, re.DOTALL)
            if matches:
                return matches[0].strip()
        
        # If no code blocks found, try to extract code-like content
        lines = response.split('\n')
        code_lines = []
        in_code = False
        
        for line in lines:
            if any(keyword in line.lower() for keyword in ['import', 'def ', 'class ', 'from ', 'if __name__']):
                in_code = True
            if in_code:
                code_lines.append(line)
        
        if code_lines:
            return '\n'.join(code_lines)
        
        return None
    
    def extract_rust_code(self, response):
        """Extract Rust code from API response - only extract content inside verus! {}"""
        # Find the start of verus! block
        verus_start = response.find('verus!')
        if verus_start == -1:
            return None
        
        # Find the opening brace after verus!
        brace_start = response.find('{', verus_start)
        if brace_start == -1:
            return None
        
        # Use bracket counting to find the matching closing brace
        bracket_count = 0
        brace_end = -1
        
        for i in range(brace_start, len(response)):
            char = response[i]
            if char == '{':
                bracket_count += 1
            elif char == '}':
                bracket_count -= 1
                if bracket_count == 0:
                    brace_end = i
                    break
        
        if brace_end == -1:
            return None
        
        # Extract content inside verus! {}
        verus_content = response[brace_start + 1:brace_end].strip()
        
        # Add the required prefix
        full_code = f"""use vstd::prelude::*;
fn main() {{}}
verus! {{
{verus_content}
}}"""
        
        return full_code
    
    def run_python_code(self, code, work_dir):
        """Run Python code and capture output"""
        try:
            # Write code to temporary file
            temp_file = work_dir / "temp_z3.py"
            with open(temp_file, 'w') as f:
                f.write(code)
            
            # Check if Z3 is available (common dependency for counterexample generation)
            z3_available = False
            try:
                import z3
                z3_available = True
            except ImportError:
                pass
            
            # Run the Python code - use relative path when using cwd
            result = subprocess.run(
                [sys.executable, "temp_z3.py"],  # Use relative path
                capture_output=True,
                text=True,
                timeout=PYTHON_EXECUTION_TIMEOUT,
                cwd=work_dir
            )
            
            # Clean up temp file
            temp_file.unlink(missing_ok=True)
            
            # If Z3 is not available and we get import error, provide helpful message
            if not z3_available and "ModuleNotFoundError: No module named 'z3'" in result.stderr:
                return "", "Z3 library not available. Install with: pip install z3-solver", 1
            
            return result.stdout, result.stderr, result.returncode
            
        except Exception as e:
            return "", f"Python execution error: {e}", 1
    
    def process_file(self, file_path, dataset_name, file_name):
        """Process a single autoverus_output.rs file"""
        print(f"Processing: {dataset_name}/{file_name}")
        
        # Create result directory structure
        result_file_dir = self.results_dir / dataset_name / file_name
        result_file_dir.mkdir(parents=True, exist_ok=True)
        
        # Copy original file
        shutil.copy2(file_path, result_file_dir / FILE_NAMES["original"])
        
        # Step 1: Initial verus verification
        start_time = time.time()
        stdout, stderr, returncode = self.run_verus(file_path)
        initial_verification_time = time.time() - start_time
        
        initial_success = self.check_verus_success(stdout, stderr)
        initial_error = stderr if not initial_success else ""
        
        print(f"  Initial verification: {'SUCCESS' if initial_success else 'FAILED'}")
        
        # If verification succeeds, we're done with this file
        if initial_success:
            self.excel_data.append([
                f"{dataset_name}/{file_name}",
                "SUCCESS",
                "",
                "N/A", "N/A", "N/A",  # z3 generation
                "N/A", "N/A", "N/A",  # z3 generation
                "N/A", "N/A", "N/A",  # z3 generation
                "N/A", "N/A", "N/A",  # z3 generation
                "N/A", "N/A", "N/A",  # z3 generation
                "N/A",  # counterexample
                "N/A", "N/A", "N/A",  # repair 1
                "N/A", "N/A", "N/A",  # repair 2
                "N/A", "N/A", "N/A",  # repair 3
                "N/A",  # repair success
                "N/A", "N/A", "N/A",  # baseline 1
                "N/A", "N/A", "N/A",  # baseline 2
                "N/A", "N/A", "N/A",  # baseline 3
                "N/A"   # baseline success
            ])
            return
        
        # Step 2: Generate z3 counterexample generator
        z3_solver_dir = result_file_dir / DIRECTORIES["z3_solver"]
        z3_solver_dir.mkdir(exist_ok=True)
        
        z3_generation_results = []
        for attempt in range(MAX_Z3_ATTEMPTS):  # Max 5 attempts
            print(f"  Generating z3 solver (attempt {attempt + 1})")
            
            prompt = PROMPTS["z3_generation"].format(
                code=open(file_path, 'r').read(),
                error=initial_error
            )
            
            start_time = time.time()
            response = self.call_deepseek_reasoner(prompt)
            api_time = time.time() - start_time
            
            if response is None:
                print(f"    API call failed")
                z3_generation_results.append(("FAILED", api_time, "API call failed"))
                break
            
            # Check if response is empty or too short
            if not response or len(response.strip()) < 10:
                print(f"    WARNING: API returned empty or very short response")
                z3_generation_results.append(("FAILED", api_time, "Empty API response"))
                continue
            
            # Extract Python code
            python_code = self.extract_python_code(response)
            if python_code is None:
                print(f"    Failed to extract Python code")
                z3_generation_results.append(("FAILED", api_time, "Failed to extract code"))
                continue
            
            # Save Python code
            python_file = z3_solver_dir / f"{attempt + 1}.py"
            with open(python_file, 'w') as f:
                f.write(python_code)
            
            # Test Python code
            start_time = time.time()
            py_stdout, py_stderr, py_returncode = self.run_python_code(python_code, z3_solver_dir)
            execution_time = time.time() - start_time
            
            if py_returncode == 0 and py_stdout.strip():
                print(f"    Z3 solver execution: SUCCESS ✓")
                print(f"    Output: {py_stdout[:100]}...")
                z3_generation_results.append(("SUCCESS", api_time, f"Output: {py_stdout[:100]}..."))
                
                # Save counterexample output
                with open(result_file_dir / FILE_NAMES["counterexample"], 'w') as f:
                    f.write(py_stdout)
                break
            else:
                print(f"    Z3 solver execution: FAILED ✗")
                error_msg = py_stderr if py_stderr else "No output"
                print(f"    Error: {error_msg[:100]}...")
                z3_generation_results.append(("FAILED", api_time, f"Error: {error_msg[:100]}..."))
                
                # Try to fix the code
                if attempt < MAX_Z3_ATTEMPTS - 1:
                    fix_prompt = PROMPTS["z3_fix"].format(
                        code=python_code,
                        error=error_msg
                    )
                    
                    fix_response = self.call_deepseek_reasoner(fix_prompt)
                    if fix_response:
                        python_code = self.extract_python_code(fix_response)
                        if python_code:
                            continue
        
        # Fill in z3 generation results
        while len(z3_generation_results) < MAX_Z3_ATTEMPTS:
            z3_generation_results.append((STATUS_VALUES["not_applicable"], STATUS_VALUES["not_applicable"], STATUS_VALUES["not_applicable"]))
        
        # Step 3: Repair based on counterexample
        repair_dir = result_file_dir / DIRECTORIES["repair_ctx"]
        repair_dir.mkdir(exist_ok=True)
        
        repair_results = []
        for attempt in range(MAX_REPAIR_ATTEMPTS):
            print(f"  Repairing code (attempt {attempt + 1})")
            
            counterexample_file = result_file_dir / FILE_NAMES["counterexample"]
            counterexample_content = ""
            if counterexample_file.exists():
                with open(counterexample_file, 'r') as f:
                    counterexample_content = f.read()
            else:
                counterexample_content = "No counterexample available"
            
            prompt = PROMPTS["repair_with_counterexample"].format(
                counterexample=counterexample_content,
                code=open(file_path, 'r').read(),
                error=initial_error
            )
            
            start_time = time.time()
            response = self.call_deepseek_reasoner(prompt)
            api_time = time.time() - start_time
            
            if response is None:
                print(f"    API call failed")
                repair_results.append(("FAILED", api_time, "API call failed"))
                continue
            
            # Check if response is empty or too short
            if not response or len(response.strip()) < 10:
                print(f"    WARNING: API returned empty or very short response")
                repair_results.append(("FAILED", api_time, "Empty API response"))
                continue
            
            # Save LM response to txt file
            response_file = repair_dir / f"{attempt + 1}.txt"
            with open(response_file, 'w') as f:
                f.write(response)
            
            # Extract Rust code
            rust_code = self.extract_rust_code(response)
            if rust_code is None:
                print(f"    Failed to extract Rust code")
                repair_results.append(("FAILED", api_time, "Failed to extract code"))
                continue
            
            # Save repaired code
            repair_file = repair_dir / f"{attempt + 1}.rs"
            with open(repair_file, 'w') as f:
                f.write(rust_code)
            
            # Verify repaired code
            start_time = time.time()
            repair_stdout, repair_stderr, repair_returncode = self.run_verus(repair_file)
            verification_time = time.time() - start_time
            
            repair_success = self.check_verus_success(repair_stdout, repair_stderr)
            
            # Output verification status to terminal
            if repair_success:
                print(f"    Verification: SUCCESS ✓")
            else:
                print(f"    Verification: FAILED ✗")
                # Show error details
                error_output = repair_stderr if repair_stderr else repair_stdout
                if error_output:
                    print(f"    Error details: {error_output[:200]}...")
            
            repair_results.append(("SUCCESS" if repair_success else "FAILED", api_time, f"Verification: {'SUCCESS' if repair_success else 'FAILED'}"))
            
            if repair_success:
                print(f"    Repair successful!")
                # Copy to parent directory as verified.rs
                shutil.copy2(repair_file, result_file_dir / FILE_NAMES["verified"])
                break
        
        # Fill in repair results
        while len(repair_results) < MAX_REPAIR_ATTEMPTS:
            repair_results.append((STATUS_VALUES["not_applicable"], STATUS_VALUES["not_applicable"], STATUS_VALUES["not_applicable"]))
        
        # Check if any repair was successful
        repair_success = any("SUCCESS" in result[2] for result in repair_results if result[2] != STATUS_VALUES["not_applicable"])
        
        # Step 4: Baseline repair (direct DeepSeek Reasoner without counterexample)
        baseline_dir = result_file_dir / DIRECTORIES["result_naive"]
        baseline_dir.mkdir(exist_ok=True)
        
        baseline_results = []
        for attempt in range(MAX_REPAIR_ATTEMPTS):
            print(f"  Baseline repair (attempt {attempt + 1})")
            
            prompt = PROMPTS["baseline_repair"].format(
                code=open(file_path, 'r').read(),
                error=initial_error
            )
            
            start_time = time.time()
            response = self.call_deepseek_reasoner(prompt)
            api_time = time.time() - start_time
            
            if response is None:
                print(f"    API call failed")
                baseline_results.append(("FAILED", api_time, "API call failed"))
                continue
            
            # Check if response is empty or too short
            if not response or len(response.strip()) < 10:
                print(f"    WARNING: API returned empty or very short response")
                baseline_results.append(("FAILED", api_time, "Empty API response"))
                continue
            
            # Save LM response to txt file
            response_file = baseline_dir / f"{attempt + 1}.txt"
            with open(response_file, 'w') as f:
                f.write(response)
            
            # Extract Rust code
            rust_code = self.extract_rust_code(response)
            if rust_code is None:
                print(f"    Failed to extract Rust code")
                baseline_results.append(("FAILED", api_time, "Failed to extract code"))
                continue
            
            # Save baseline repaired code
            baseline_file = baseline_dir / f"{attempt + 1}.rs"
            with open(baseline_file, 'w') as f:
                f.write(rust_code)
            
            # Verify baseline repaired code
            start_time = time.time()
            baseline_stdout, baseline_stderr, baseline_returncode = self.run_verus(baseline_file)
            verification_time = time.time() - start_time
            
            baseline_success = self.check_verus_success(baseline_stdout, baseline_stderr)
            
            # Output verification status to terminal
            if baseline_success:
                print(f"    Verification: SUCCESS ✓")
            else:
                print(f"    Verification: FAILED ✗")
                # Show error details
                error_output = baseline_stderr if baseline_stderr else baseline_stdout
                if error_output:
                    print(f"    Error details: {error_output[:200]}...")
            
            baseline_results.append(("SUCCESS" if baseline_success else "FAILED", api_time, f"Verification: {'SUCCESS' if baseline_success else 'FAILED'}"))
            
            if baseline_success:
                print(f"    Baseline repair successful!")
                break
        
        # Fill in baseline results
        while len(baseline_results) < MAX_REPAIR_ATTEMPTS:
            baseline_results.append((STATUS_VALUES["not_applicable"], STATUS_VALUES["not_applicable"], STATUS_VALUES["not_applicable"]))
        
        # Check if any baseline repair was successful
        baseline_success = any("SUCCESS" in result[2] for result in baseline_results if result[2] != STATUS_VALUES["not_applicable"])
        
        # Add to excel data
        counterexample_content = ""
        counterexample_file = result_file_dir / FILE_NAMES["counterexample"]
        if counterexample_file.exists():
            with open(counterexample_file, 'r') as f:
                content = f.read()
                counterexample_content = content[:500] + "..." if len(content) > 500 else content
        
        row = [
            f"{dataset_name}/{file_name}",
            STATUS_VALUES["failed"],
            initial_error[:200] + "..." if len(initial_error) > 200 else initial_error
        ]
        
        # Add z3 generation results
        for result in z3_generation_results:
            row.extend([result[0], result[1], result[2]])
        
        # Add counterexample
        row.append(counterexample_content)
        
        # Add repair results
        for result in repair_results:
            row.extend([result[0], result[1], result[2]])
        
        # Add repair success
        row.append(STATUS_VALUES["success"] if repair_success else STATUS_VALUES["failed"])
        
        # Add baseline results
        for result in baseline_results:
            row.extend([result[0], result[1], result[2]])
        
        # Add baseline success
        row.append(STATUS_VALUES["success"] if baseline_success else STATUS_VALUES["failed"])
        
        self.excel_data.append(row)
    
    def run_experiment(self):
        """Run the complete experiment"""
        print(f"Starting experiment at {datetime.now()}")
        print(f"Results will be saved to: {self.results_dir}")
        
        # Process all files
        for dataset_dir in self.base_dir.iterdir():
            if dataset_dir.is_dir():
                dataset_name = dataset_dir.name
                print(f"\nProcessing dataset: {dataset_name}")
                
                for file_dir in dataset_dir.iterdir():
                    if file_dir.is_dir():
                        file_name = file_dir.name
                        autoverus_file = file_dir / "autoverus_output.rs"
                        
                        if autoverus_file.exists():
                            try:
                                self.process_file(autoverus_file, dataset_name, file_name)
                            except Exception as e:
                                print(f"Error processing {dataset_name}/{file_name}: {e}")
                                # Add error row to excel data
                                self.excel_data.append([
                                    f"{dataset_name}/{file_name}",
                                    "ERROR",
                                    str(e),
                                    "ERROR", "ERROR", "ERROR",
                                    "ERROR", "ERROR", "ERROR",
                                    "ERROR", "ERROR", "ERROR",
                                    "ERROR", "ERROR", "ERROR",
                                    "ERROR", "ERROR", "ERROR",
                                    "ERROR",
                                    "ERROR", "ERROR", "ERROR",
                                    "ERROR", "ERROR", "ERROR",
                                    "ERROR", "ERROR", "ERROR",
                                    "ERROR",
                                    "ERROR", "ERROR", "ERROR",
                                    "ERROR", "ERROR", "ERROR",
                                    "ERROR", "ERROR", "ERROR",
                                    "ERROR"
                                ])
        
        # Generate Excel report
        self.generate_excel_report()
        
        # Generate summary tables
        self.generate_summary_tables()
        
        print(f"\nExperiment completed! Results saved to: {self.results_dir}")
    
    def generate_excel_report(self):
        """Generate detailed Excel report"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Detailed Results"
        
        # Headers
        headers = EXCEL_HEADERS["detailed"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Data
        for row_idx, row_data in enumerate(self.excel_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        excel_file = self.results_dir / EXCEL_FILES["detailed"]
        wb.save(excel_file)
        print(f"Detailed Excel report saved to: {excel_file}")
    
    def generate_summary_tables(self):
        """Generate summary tables"""
        wb = openpyxl.Workbook()
        
        # Summary table 1: File-level results
        ws1 = wb.active
        ws1.title = "File Results Summary"
        
        headers1 = EXCEL_HEADERS["file_summary"]
        
        for col, header in enumerate(headers1, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Process data for summary
        for row_idx, row_data in enumerate(self.excel_data, 2):
            folder_name = row_data[0]
            initial_success = row_data[1]
            counterexample_generated = "YES" if row_data[18] != "N/A" and row_data[18] != "ERROR" else "NO"
            repair_success = row_data[28]
            baseline_success = row_data[38]
            
            ws1.cell(row=row_idx, column=1, value=folder_name)
            ws1.cell(row=row_idx, column=2, value=initial_success)
            ws1.cell(row=row_idx, column=3, value=counterexample_generated)
            ws1.cell(row=row_idx, column=4, value=repair_success)
            ws1.cell(row=row_idx, column=5, value=baseline_success)
        
        # Auto-adjust column widths
        for column in ws1.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws1.column_dimensions[column_letter].width = adjusted_width
        
        # Summary table 2: Dataset overview
        ws2 = wb.create_sheet("Dataset Overview")
        
        headers2 = EXCEL_HEADERS["dataset_overview"]
        
        for col, header in enumerate(headers2, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Calculate dataset statistics
        dataset_stats = {}
        for row_data in self.excel_data:
            dataset_name = row_data[0].split('/')[0]
            if dataset_name not in dataset_stats:
                dataset_stats[dataset_name] = {
                    'total': 0,
                    'initially_verified': 0,
                    'counterexample_repaired': 0,
                    'direct_repaired': 0
                }
            
            dataset_stats[dataset_name]['total'] += 1
            
            if row_data[1] == "SUCCESS":
                dataset_stats[dataset_name]['initially_verified'] += 1
            
            if row_data[28] == "SUCCESS":
                dataset_stats[dataset_name]['counterexample_repaired'] += 1
            
            if row_data[38] == "SUCCESS":
                dataset_stats[dataset_name]['direct_repaired'] += 1
        
        # Add dataset statistics to sheet
        for row_idx, (dataset, stats) in enumerate(dataset_stats.items(), 2):
            ws2.cell(row=row_idx, column=1, value=dataset)
            ws2.cell(row=row_idx, column=2, value=stats['total'])
            ws2.cell(row=row_idx, column=3, value=stats['initially_verified'])
            ws2.cell(row=row_idx, column=4, value=stats['counterexample_repaired'])
            ws2.cell(row=row_idx, column=5, value=stats['direct_repaired'])
        
        # Auto-adjust column widths
        for column in ws2.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws2.column_dimensions[column_letter].width = adjusted_width
        
        # Save summary workbook
        summary_file = self.results_dir / EXCEL_FILES["summary"]
        wb.save(summary_file)
        print(f"Summary tables saved to: {summary_file}")

def main():
    # Use DeepSeek API key from config
    api_key = API_KEY
    
    # Create and run experiment
    runner = ExperimentRunner(api_key)
    runner.run_experiment()

if __name__ == "__main__":
    main()
